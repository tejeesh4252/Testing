"""
Import/export service: bank-file -> cashbook import, cashbook GL write-back,
month archiving (including the auto-archive that was missing from v3.5),
and Yardi ETL export.

Two dedupe layers exist here and they check different things:
  - Cashbook-level (raw bank import): is this bank row already a row in
    the cashbook sheet? Keyed on (date, amount, description) -- not just
    (date, amount) -- so two legitimate same-day, same-amount
    transactions (a recurring transfer, a duplicate wire fee) are no
    longer silently collapsed into one and dropped.
  - DB-level (archive): handled in db.Database.insert_transactions,
    keyed on (fund, date, amount, description) as well.

Auto-archive: this is the feature that was described but never actually
wired up in the original script. After a bank import appends new rows,
archive_completed_months() scans the cashbook for any month strictly
before the month that was just imported, and archives it to the database
automatically -- but only if every row in that month already has a GL
mapping filled in. A month with unmapped rows is left alone and flagged
in the log, so nothing gets archived half-mapped.
"""

from __future__ import annotations

import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from db import Database

LogFn = Callable[[str], None]


def _noop_log(_msg: str) -> None:
    pass


# ============================================================================
# TRANSACTION READER -- reads a monthly cashbook workbook sheet-by-sheet,
# used by the review/auto-mapping workflow (as opposed to the raw bank
# import workflow below, which writes directly into the cashbook).
# ============================================================================

class TransactionReader:

    @staticmethod
    def _clean(val) -> str:
        s = str(val).strip()
        return "" if s in ("nan", "None", "NaN", " ", "") else s

    @staticmethod
    def _is_gl_code(val) -> bool:
        s = str(val).strip()
        return bool(s and s not in ("nan", "None") and any(c.isdigit() for c in s))

    def find_header_row(self, df_raw) -> Optional[int]:
        for idx, row in df_raw.iterrows():
            vals = [str(v).strip().lower() for v in row.values]
            if any(k in vals for k in ["post date", "transaction date", "date"]):
                return idx
        return None

    def get_bank_gl(self, df_raw) -> Optional[str]:
        limit = min(10, len(df_raw))
        for idx in range(limit):
            row = df_raw.iloc[idx]
            for col_idx, val in enumerate(row.values):
                if str(val).strip().lower() == "gl - code":
                    try:
                        gl = str(row.values[col_idx + 1]).strip()
                        if self._is_gl_code(gl):
                            return gl
                    except IndexError:
                        pass
        return None

    @staticmethod
    def build_combined(bank_desc: str, bank_desc2: str,
                        manual_desc: str, je_comment: str) -> str:
        bd, bd2 = bank_desc.strip(), bank_desc2.strip()
        md, je = manual_desc.strip(), je_comment.strip()
        if je:
            return je
        if md:
            if " - " in md:
                return md
            if bd and bd2:
                return f"{bd} - {bd2}"
            return md
        if bd and bd2:
            return f"{bd} - {bd2}"
        return bd or bd2 or ""

    def read_sheet(self, filepath, sheet_name: str, log: LogFn = _noop_log):
        try:
            df_raw = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        except Exception as e:
            log(f"  Cannot open '{sheet_name}': {e}")
            return [], None

        bank_gl = self.get_bank_gl(df_raw)
        header_idx = self.find_header_row(df_raw)
        if header_idx is None:
            return [], bank_gl

        df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_idx, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        def gcol(*names):
            for n in names:
                if n in df.columns:
                    return n
            return None

        col_date = gcol("Post Date", "Transaction Date", "Date", "post date")
        col_amt = gcol("Transaction Amount", "Amount", "Debit", "Credit")
        col_bd = gcol("Transaction Description", "Bank Description", "Description", "Narration")
        col_bd2 = gcol("Transaction Description 2", "Description 2")
        col_md = gcol("Manual Description", "Manual Desc")
        col_je = gcol("JE comments", "JE Comments", "Comments")
        col_exgl = gcol("Yardi Account #", "GL Code", "GL Account")
        col_exname = gcol("Yardi Account name", "GL Name", "Account Name")

        if not col_date:
            log(f"  No date column in '{sheet_name}'")
            return [], bank_gl

        transactions = []
        skipped_date = skipped_amt = 0

        for _, row in df.iterrows():
            raw_date = self._clean(row.get(col_date, ""))
            if not raw_date:
                skipped_date += 1
                continue
            try:
                date_num = float(raw_date)
                if date_num < 1000:
                    skipped_date += 1
                    continue
            except (ValueError, TypeError):
                try:
                    dt = pd.to_datetime(raw_date, dayfirst=False)
                    date_num = (dt - pd.Timestamp("1899-12-30")).days
                except Exception:
                    skipped_date += 1
                    continue

            raw_amt = self._clean(row.get(col_amt, "") if col_amt else "")
            try:
                amount = float(str(raw_amt).replace(",", "").replace("(", "-").replace(")", ""))
            except (ValueError, TypeError):
                skipped_amt += 1
                continue

            bd = self._clean(row.get(col_bd, "") if col_bd else "")
            bd2 = self._clean(row.get(col_bd2, "") if col_bd2 else "")
            md = self._clean(row.get(col_md, "") if col_md else "")
            je = self._clean(row.get(col_je, "") if col_je else "")
            combined = self.build_combined(bd, bd2, md, je)

            existing_gl = self._clean(row.get(col_exgl, "") if col_exgl else "")
            existing_nm = self._clean(row.get(col_exname, "") if col_exname else "")
            if existing_gl and not self._is_gl_code(existing_gl):
                existing_gl, existing_nm = "", ""

            transactions.append({
                "date": date_num, "amount": amount,
                "bank_desc": bd, "bank_desc2": bd2, "combined_desc": combined,
                "comments": je, "existing_gl": existing_gl or None,
                "existing_name": existing_nm or None,
                "bank_gl": bank_gl, "sheet": sheet_name,
            })

        log(f"  {len(transactions)} transactions extracted "
            f"(skipped: {skipped_date} bad dates, {skipped_amt} bad amounts)")
        return transactions, bank_gl


# ============================================================================
# ETL EXPORTER -- writes the Yardi journal-entry import file
# ============================================================================

class ETLExporter:

    def export(self, mapped_transactions, fund_name: str, property_code: str,
               book_num: int, output_path: str):
        rows = []
        transactions_by_date = defaultdict(list)

        for txn in mapped_transactions:
            if txn.get("status") == "SKIP":
                continue
            if not txn.get("mapped_gl") or not txn.get("bank_gl"):
                continue

            date_val = txn["date"]
            try:
                dt = pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(date_val))
                date_str = dt.strftime("%m/%d/%Y")
                post_mon = dt.replace(day=1).strftime("%Y-%m")
                date_key = dt.date()
            except Exception:
                date_str, post_mon, date_key = str(date_val), "", date_val

            transactions_by_date[date_key].append({
                "date_str": date_str, "post_mon": post_mon,
                "amount": txn["amount"], "bank_gl": txn["bank_gl"],
                "offset_gl": txn["mapped_gl"],
                "detail_note": txn.get("combined_desc", ""),
            })

        if not transactions_by_date:
            return False, "No transactions to export"

        sorted_dates = sorted(transactions_by_date.keys())
        for tran_num, date_key in enumerate(sorted_dates, start=1):
            for txn in transactions_by_date[date_key]:
                rows.append([tran_num, txn["date_str"], property_code, txn["bank_gl"],
                             txn["post_mon"], book_num, txn["amount"], txn["detail_note"], "", ""])
                rows.append([tran_num, txn["date_str"], property_code, txn["offset_gl"],
                             txn["post_mon"], book_num, -txn["amount"], txn["detail_note"], "", ""])

        df_out = pd.DataFrame(rows, columns=[
            "TRANNUM", "DATE", "PROPERTY", "ACCOUNT", "POSTMONTH",
            "BOOKNUM", "AMOUNT", "REMARK", "REF", "DETAILNOTES"
        ])

        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                pd.DataFrame([["FinJournals"]]).to_excel(
                    writer, sheet_name="ETL Export", index=False, header=False)
                df_out.to_excel(writer, sheet_name="ETL Export", index=False, startrow=1)

            wb = openpyxl.load_workbook(output_path)
            ws = wb["ETL Export"]

            banner_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            banner_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
            for cell in ws[1]:
                cell.fill = banner_fill
                cell.font = banner_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
            for cell in ws[2]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    if cell.column == 5 and cell.value:
                        try:
                            dt = datetime.strptime(str(cell.value), "%Y-%m")
                            cell.value = dt
                            cell.number_format = "MM/DD/YYYY"
                            cell.alignment = Alignment(horizontal="center")
                        except Exception:
                            pass
                    if cell.column == 2 and cell.value:
                        cell.alignment = Alignment(horizontal="center")
                    if cell.column == 7 and cell.value not in ("", None):
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right")

            col_widths = {1: 10, 2: 12, 3: 8, 4: 14, 5: 14, 6: 10, 7: 14, 8: 35, 9: 10, 10: 10}
            for col_idx, width in col_widths.items():
                ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

            wb.save(output_path)
            wb.close()
            return True, f"Exported {len(sorted_dates)} journal entries ({len(rows) // 2} transactions)"
        except Exception as e:
            return False, str(e)


# ============================================================================
# RAW BANK IMPORT -- multi-account bank statement -> cashbook columns A-D.
# No GL mapping happens here; that's the review/auto-mapping workflow above.
# ============================================================================

def _convert_to_datetime(value):
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y", "%m-%d-%Y"]:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    try:
        num_val = float(value)
        if 40000 <= num_val <= 60000:  # realistic Excel date range
            return pd.to_datetime("1899-12-30") + pd.Timedelta(days=int(num_val))
    except Exception:
        pass
    return None


def _get_date_range_from_cashbook(cashbook_path, sheet_mapping, header_rows):
    try:
        wb = openpyxl.load_workbook(cashbook_path, data_only=True)
        dates = []
        for sn in sheet_mapping.values():
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            for r in range(header_rows + 1, ws.max_row + 1):
                cell = ws.cell(r, 1)
                if cell.value is None:
                    break
                if isinstance(cell.value, datetime):
                    dates.append(cell.value)
        wb.close()
        return (min(dates), max(dates)) if dates else (None, None)
    except Exception:
        return None, None


def _is_valid_transaction_row(row, exclude_keywords):
    date_val = row.iloc[1] if len(row) > 1 else None
    amount_val = row.iloc[2] if len(row) > 2 else None
    desc_val = str(row.iloc[4]).strip().upper() if len(row) > 4 and pd.notna(row.iloc[4]) else ""

    if not _convert_to_datetime(date_val):
        return False, "No valid date"
    for keyword in exclude_keywords:
        if keyword in desc_val:
            return False, f"Excluded keyword: {keyword}"
    if amount_val is None or str(amount_val).strip() in ("", "nan", "None"):
        return False, "No amount"
    try:
        float(amount_val)
    except (ValueError, TypeError):
        return False, "Amount not numeric"
    return True, "Valid"


def _get_existing_transactions(ws, header_rows):
    """Read (date, amount, description) triples for duplicate detection.
    Description is included so two legitimate same-day, same-amount
    transactions are not treated as duplicates of each other."""
    existing = set()
    for r in range(header_rows + 1, ws.max_row + 1):
        date_cell = ws.cell(r, 1).value
        amount_cell = ws.cell(r, 2).value
        desc_cell = ws.cell(r, 3).value
        if date_cell is None:
            break
        if isinstance(date_cell, datetime) and amount_cell is not None:
            try:
                existing.add((date_cell.date(), round(float(amount_cell), 2),
                              str(desc_cell or "").strip().lower()))
            except (TypeError, ValueError):
                pass
    return existing


def _identify_account_sections(df, bank_normalization):
    accounts, current_account = [], None
    for idx, row in df.iterrows():
        col1 = str(row.iloc[1]).strip().upper() if pd.notna(row.iloc[1]) and len(row) > 1 else ""
        col2 = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) and len(row) > 2 else ""
        col3 = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) and len(row) > 3 else ""

        bank_norm = None
        for key, val in bank_normalization.items():
            if key in col1:
                bank_norm = val
                break

        try:
            col2_clean = str(int(float(col2))).replace("-", "").replace(" ", "")
            col2_str = str(int(float(col2)))
        except (ValueError, TypeError):
            col2_clean = col2.replace("-", "").replace(" ", "")
            col2_str = col2

        is_account_number = col2_clean.isdigit() and len(col2_clean) >= 6

        if bank_norm and is_account_number:
            if current_account and current_account["end_row"] is None:
                current_account["end_row"] = idx - 1
            current_account = {
                "bank_name": bank_norm, "account_number": col2_str,
                "account_type": col3, "start_row": idx + 1, "end_row": None,
            }
            accounts.append(current_account)

    if current_account and current_account["end_row"] is None:
        current_account["end_row"] = len(df) - 1
    return accounts


def _extract_account_data(df, account_info, exclude_keywords, log: LogFn):
    start = account_info["start_row"]
    end = account_info["end_row"] if account_info["end_row"] is not None else len(df) - 1

    sub = df.iloc[start:end + 1].copy().dropna(how="all")
    valid, excl = [], 0
    for _, row in sub.iterrows():
        ok, _ = _is_valid_transaction_row(row, exclude_keywords)
        if ok:
            valid.append(row)
        else:
            excl += 1

    log(f"  Valid: {len(valid)}, excluded: {excl}")
    if not valid:
        return pd.DataFrame()

    tdf = pd.DataFrame(valid)
    keep = [i for i in range(len(tdf.columns)) if i not in [0, 4]]
    tdf = tdf.iloc[:, keep]
    if not tdf.empty:
        tdf.iloc[:, 0] = tdf.iloc[:, 0].apply(_convert_to_datetime)
        tdf = tdf.sort_values(by=tdf.columns[0]).reset_index(drop=True)
    return tdf


def _append_to_cashbook(transaction_df, sheet_name, cashbook_path, header_rows, log: LogFn):
    """Append valid rows to cashbook columns A-D. Returns (added, skipped).
    Takes a backup of the cashbook before writing."""
    backup_path = None
    try:
        p = Path(cashbook_path)
        backup_path = p.with_name(f"{p.stem}.bak{p.suffix}")
        shutil.copy2(cashbook_path, backup_path)
    except Exception as e:
        log(f"  Warning: could not create cashbook backup ({e})")

    wb = openpyxl.load_workbook(cashbook_path)
    if sheet_name not in wb.sheetnames:
        log(f"  Sheet '{sheet_name}' not found!")
        wb.close()
        return 0, 0

    ws = wb[sheet_name]
    existing = _get_existing_transactions(ws, header_rows)
    log(f"  Existing transactions on sheet: {len(existing)}")

    next_row = header_rows + 1
    while ws.cell(next_row, 1).value is not None:
        next_row += 1

    added = skipped = 0
    for _, row in transaction_df.iterrows():
        trans_date = row.iloc[0]
        amount = float(row.iloc[1])
        desc = str(row.iloc[2]).strip().lower() if len(row) > 2 and pd.notna(row.iloc[2]) else ""
        key = (trans_date.date(), round(amount, 2), desc)

        if key in existing:
            skipped += 1
            continue

        ws.cell(next_row, 1).value = trans_date
        ws.cell(next_row, 2).value = amount
        if len(row) > 2 and pd.notna(row.iloc[2]):
            ws.cell(next_row, 3).value = str(row.iloc[2])
        if len(row) > 3 and pd.notna(row.iloc[3]):
            ws.cell(next_row, 4).value = str(row.iloc[3])

        added += 1
        next_row += 1
        existing.add(key)

    wb.save(cashbook_path)
    wb.close()
    log(f"  Added: {added} | Skipped (dups): {skipped}")
    return added, skipped


class BankImporter:
    """Reads a raw bank statement file, identifies per-account sections,
    and appends new rows to the matching cashbook sheet."""

    def __init__(self, header_rows: int, exclude_keywords: list):
        self.header_rows = header_rows
        self.exclude_keywords = exclude_keywords

    def run(self, source_path: str, cashbook_path: str, fund_cfg: dict, log: LogFn = _noop_log):
        log(f"Reading source file: {Path(source_path).name}")
        df = pd.read_excel(source_path, sheet_name=0, header=None)
        log(f"Loaded {len(df)} rows, {len(df.columns)} columns")

        c_early, c_late = _get_date_range_from_cashbook(
            cashbook_path, fund_cfg["sheet_mapping"], self.header_rows)
        if c_early:
            log(f"Cashbook date range: {c_early.date()} to {c_late.date()}")

        accounts = _identify_account_sections(df, fund_cfg["bank_normalization"])
        if not accounts:
            log("No account sections found in source file.")
            return {"total_added": 0, "total_skipped": 0, "latest_month": None, "accounts": []}

        log(f"Found {len(accounts)} account section(s)")
        for acc in accounts:
            key = f"{acc['bank_name']}_{acc['account_number'].replace('-', '').replace(' ', '')}"
            acc["sheet_name"] = fund_cfg["sheet_mapping"].get(key, key)
            log(f"  {acc['bank_name']} {acc['account_number']} -> {acc['sheet_name']}")

        total_added = total_skipped = 0
        latest_month = None

        for acc in accounts:
            log(f"-- {acc['bank_name']} {acc['account_number']} --")
            data = _extract_account_data(df, acc, self.exclude_keywords, log)
            if data.empty:
                log("  No data found")
                continue

            for d in data.iloc[:, 0]:
                m = d.strftime("%Y-%m")
                if latest_month is None or m > latest_month:
                    latest_month = m

            added, skipped = _append_to_cashbook(
                data, acc["sheet_name"], cashbook_path, self.header_rows, log)
            total_added += added
            total_skipped += skipped

        return {
            "total_added": total_added, "total_skipped": total_skipped,
            "latest_month": latest_month, "accounts": accounts,
        }


# ============================================================================
# CASHBOOK GL WRITE-BACK -- writes mapped GL codes into cashbook columns E-F
# ============================================================================

class CashbookGLWriter:

    def save(self, cashbook_path: str, mapped_transactions, header_rows: int, log: LogFn = _noop_log):
        backup_path = None
        try:
            p = Path(cashbook_path)
            backup_path = p.with_name(f"{p.stem}.bak{p.suffix}")
            shutil.copy2(cashbook_path, backup_path)
        except Exception as e:
            log(f"Warning: could not create cashbook backup ({e})")

        wb = openpyxl.load_workbook(cashbook_path)
        updated = 0
        for txn in mapped_transactions:
            sheet_name = txn.get("sheet")
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            try:
                txn_date = (pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(txn["date"]))).date()
            except Exception:
                continue
            amount = txn["amount"]

            for r in range(header_rows + 1, ws.max_row + 1):
                cell_date = ws.cell(r, 1).value
                cell_amount = ws.cell(r, 2).value
                if cell_date is None:
                    break
                if not isinstance(cell_date, datetime):
                    continue
                try:
                    row_amount = float(cell_amount)
                except (TypeError, ValueError):
                    continue
                if cell_date.date() == txn_date and abs(row_amount - amount) < 0.01:
                    ws.cell(r, 6).value = txn.get("mapped_gl", "")
                    ws.cell(r, 7).value = txn.get("mapped_name", "")
                    updated += 1
                    break

        wb.save(cashbook_path)
        wb.close()
        return updated, backup_path


# ============================================================================
# MONTH ARCHIVER -- the connective tissue that was missing from v3.5:
# after an import, automatically archive any *closed* prior month whose
# rows are already fully GL-mapped. A month with any unmapped row is
# left alone and reported, not partially archived.
# ============================================================================

class MonthArchiver:

    def __init__(self, db: Database, header_rows: int):
        self.db = db
        self.header_rows = header_rows

    def archive_reviewed_batch(self, transactions, fund_name: str,
                                property_code: str, source_month: str):
        """Archive transactions that came out of the manual review /
        auto-mapping workflow (already carry mapped_gl)."""
        self.db.backup(tag=f"archive_{fund_name}")
        mapped = [t for t in transactions if t.get("mapped_gl") and t.get("status") != "SKIP"]
        return self.db.insert_transactions(fund_name, property_code, source_month, mapped)

    def archive_completed_months(self, cashbook_path: str, fund_name: str,
                                  property_code: str, sheet_mapping: dict,
                                  just_imported_month: Optional[str], log: LogFn = _noop_log):
        """Scan every mapped sheet for months strictly before
        just_imported_month. Archive a month only if every row in it
        already has a GL code in column F and it hasn't been archived
        for this fund yet."""
        if not just_imported_month:
            return {"archived_months": [], "pending_months": []}

        already_archived = self.db.archived_months(fund_name)
        wb = openpyxl.load_workbook(cashbook_path, data_only=True)

        by_month: dict = defaultdict(list)
        for sheet_name in sheet_mapping.values():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for r in range(self.header_rows + 1, ws.max_row + 1):
                date_cell = ws.cell(r, 1).value
                if date_cell is None:
                    break
                if not isinstance(date_cell, datetime):
                    continue
                month = date_cell.strftime("%Y-%m")
                if month >= just_imported_month:
                    continue  # only closed prior months are auto-archived
                gl_code = ws.cell(r, 6).value
                gl_name = ws.cell(r, 7).value
                amount = ws.cell(r, 2).value
                desc = ws.cell(r, 3).value
                desc2 = ws.cell(r, 4).value
                try:
                    amount = float(amount)
                except (TypeError, ValueError):
                    continue
                by_month[month].append({
                    "date": (date_cell - datetime(1899, 12, 30)).days,
                    "amount": amount,
                    "bank_desc": str(desc or ""),
                    "bank_desc2": str(desc2 or ""),
                    "combined_desc": str(desc or ""),
                    "mapped_gl": str(gl_code or ""),
                    "mapped_name": str(gl_name or ""),
                    "bank_gl": "",
                    "sheet": sheet_name,
                    "status": "MAPPED" if gl_code else "UNMAPPED",
                })
        wb.close()

        archived_months, pending_months = [], []
        self.db.backup(tag=f"autoarchive_{fund_name}")

        for month, rows in sorted(by_month.items()):
            if month in already_archived:
                continue
            unmapped = [r for r in rows if not r["mapped_gl"]]
            if unmapped:
                pending_months.append((month, len(unmapped)))
                log(f"  {month}: {len(unmapped)} row(s) still unmapped -- skipped, needs manual review")
                continue
            inserted, skipped = self.db.insert_transactions(fund_name, property_code, month, rows)
            archived_months.append((month, inserted, skipped))
            log(f"  {month}: archived {inserted} row(s), {skipped} already present")

        return {"archived_months": archived_months, "pending_months": pending_months}
